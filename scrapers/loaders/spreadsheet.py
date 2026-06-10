from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator, Optional

from openpyxl import load_workbook

CATEGORY_HEADERS = {
    "hardware": "hardware",
    "software": "software",
    "founders office": "founders_office",
    "product gtm": "product_gtm",
}


@dataclass
class SpreadsheetEvent:
    event_name: str
    host: str = ""
    target_roles: str = ""
    sourcing_strategy: str = ""
    timeline: str = ""
    notes: str = ""
    url: Optional[str] = None
    category: str = "software"


def _normalise_header(value: object) -> str:
    return str(value or "").strip().lower()


def _detect_category(section_title: str) -> str:
    lowered = section_title.lower()
    for key, category in CATEGORY_HEADERS.items():
        if key in lowered:
            return category
    return "software"


def load_events_from_spreadsheet(path: Path) -> list[SpreadsheetEvent]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    events: list[SpreadsheetEvent] = []
    current_category = "software"

    header_map: dict[str, int] = {}
    url_col: Optional[int] = None

    for row in sheet.iter_rows(values_only=True):
        if not row or all(cell is None for cell in row):
            continue

        first = str(row[0] or "").strip()
        if first in CATEGORY_HEADERS or any(k in first.lower() for k in CATEGORY_HEADERS):
            current_category = _detect_category(first)
            header_map = {}
            url_col = None
            continue

        if _normalise_header(row[1] if len(row) > 1 else "") == "event name":
            header_map = {
                _normalise_header(value): idx
                for idx, value in enumerate(row)
                if value is not None
            }
            if "url" in header_map:
                url_col = header_map["url"]
            elif "scrape url" in header_map:
                url_col = header_map["scrape url"]
            continue

        if not header_map:
            continue

        def cell(name: str, fallback_idx: int) -> str:
            idx = header_map.get(name, fallback_idx)
            if idx >= len(row):
                return ""
            return str(row[idx] or "").strip()

        event_name = cell("event name", 1)
        if not event_name:
            continue

        number_cell = str(row[0] or "").strip()
        if not number_cell.replace(".", "", 1).isdigit():
            continue

        scrape_url = None
        if url_col is not None and url_col < len(row) and row[url_col]:
            scrape_url = str(row[url_col]).strip()

        events.append(
            SpreadsheetEvent(
                event_name=event_name,
                host=cell("host / organizer", 2),
                target_roles=cell("target roles", 3),
                sourcing_strategy=cell("sourcing strategy", 4),
                timeline=cell("typical timeline", 5),
                notes=cell("notes", 6),
                url=scrape_url,
                category=current_category,
            ),
        )

    workbook.close()
    return events


def iter_scrape_targets(path: Path):
    from config.platform_registry import resolve_targets_for_event

    for event in load_events_from_spreadsheet(path):
        for target in resolve_targets_for_event(
            event.event_name,
            url=event.url,
            host=event.host,
            notes=event.notes,
            category_hint=event.category,
        ):
            yield target
